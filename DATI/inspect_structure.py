import openpyxl
import os

base = r"c:\Users\10294484\Desktop\AI DEMO\ALARM MANAGER\DATI"

# Inspect active alarms file
print("=== ACTIVE ALARMS FILE ===")
wb = openpyxl.load_workbook(os.path.join(base, "fm-active-Alarm Monitor-20052026.xlsx"), read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}, rows: {ws.max_row}, cols: {ws.max_column}")
    # print first 3 rows
    count = 0
    for row in ws.iter_rows(values_only=True):
        if row[0] is None and all(c is None for c in row):
            continue
        print(f"  Row {count}: {row}")
        count += 1
        if count >= 5:
            break
wb.close()

print()
print("=== HISTORY FILE (first file) ===")
hist_dir = os.path.join(base, "fm-history-20daysALL")
first_file = os.path.join(hist_dir, "fm-history-20daysALL-1.xlsx")
wb2 = openpyxl.load_workbook(first_file, read_only=True, data_only=True)
for sheet_name in wb2.sheetnames:
    ws2 = wb2[sheet_name]
    print(f"Sheet: {sheet_name}, rows: {ws2.max_row}, cols: {ws2.max_column}")
    count = 0
    for row in ws2.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        print(f"  Row {count}: {row}")
        count += 1
        if count >= 5:
            break
wb2.close()

# Also check fm-history-20daysALL.xlsx (the main one)
print()
print("=== HISTORY MAIN FILE ===")
main_hist = os.path.join(hist_dir, "fm-history-20daysALL.xlsx")
wb3 = openpyxl.load_workbook(main_hist, read_only=True, data_only=True)
for sheet_name in wb3.sheetnames:
    ws3 = wb3[sheet_name]
    print(f"Sheet: {sheet_name}, rows: {ws3.max_row}, cols: {ws3.max_column}")
    count = 0
    for row in ws3.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        print(f"  Row {count}: {row}")
        count += 1
        if count >= 5:
            break
wb3.close()
print("DONE")
