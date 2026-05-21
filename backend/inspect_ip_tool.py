"""
Read and inspect the TOOL CALCOLO IP ZTE.xlsx file to understand IP logic.
"""
import openpyxl
import os

path = r"c:\Users\10294484\Desktop\AI DEMO\ALARM MANAGER\TOOL CALCOLO IP ZTE.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== SHEET: {sheet_name} (rows={ws.max_row}, cols={ws.max_column}) ===")
    rows_printed = 0
    for row in ws.iter_rows(values_only=True):
        if all(c is None for c in row):
            continue
        print(f"  {row}")
        rows_printed += 1
        if rows_printed >= 60:
            print("  ... (truncated)")
            break
    print()

wb.close()
